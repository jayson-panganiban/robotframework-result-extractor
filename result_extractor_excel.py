import os
import fnmatch
import datetime
from lxml import etree
from openpyxl import Workbook
from itertools import count
try:
    # Python 2
    from itertools import izip
except ImportError:
    # Python 3
    izip = zip


def save_test_result_to_excel_file():
    tc_name_list = []
    tc_status_list = []
    tc_error_list = []
    tc_tag_list = []
    xml_files = get_all_xml_files()
    for each_xml_file in xml_files:
        context = etree.iterparse(each_xml_file, events=('end',))
        for event, elem in context:
            if elem.tag == 'test':
                tc_name_attrib = elem.attrib['name']
                tc_name_list.append(tc_name_attrib)
                tc_status_list.append(get_test_status_path(
                    elem, tc_name_attrib)[0].attrib['status'])
                tc_error_list.append(get_test_status_path(elem, tc_name_attrib)[0].text)
                tc_tag_list.append(get_test_tags(elem, tc_name_attrib))
                elem.clear()
                while elem.getprevious() is not None:
                    del elem.getparent()[0]
        del context

    date_stamp = '{:%Y-%m-%d-%H%M%S}'.format(datetime.datetime.now())
    result_file = 'test_result-' + date_stamp + '.xlsx'
    wb = Workbook()
    ws = wb.worksheets[0]
    ws['A1'] = 'Test Case'
    ws['B1'] = 'Status'
    ws['C1'] = 'Error Details'
    ws['D1'] = 'Tags'
    ws['E1'] = 'Comments'
    for i, name, status, error, tags in izip(count(), tc_name_list, tc_status_list, tc_error_list, tc_tag_list):
        ws['A' + str(i + 2)] = name
        ws['B' + str(i + 2)] = status
        ws['C' + str(i + 2)] = error
        ws['D' + str(i + 2)] = ', '.join(tags)
    ws.auto_filter.ref = 'A1:E1'
    wb.save(filename=result_file)


def get_test_status_path(elem, tc_name_attrib):
    if "'" in tc_name_attrib:
        tc_name_quoted = '"%s"' % tc_name_attrib
        return elem.xpath("//test[@name=" + tc_name_quoted + "]/status")
    else:
        return elem.xpath("//test[@name='" + tc_name_attrib + "']/status")


def get_test_tags(elem, tc_name_attrib):
    if "'" in tc_name_attrib:
        tc_name_quoted = '"%s"' % tc_name_attrib
        return elem.xpath("//test[@name=" + tc_name_quoted + "]/tags/tag/text()")
    else:
        return elem.xpath("//test[@name='" + tc_name_attrib + "']/tags/tag/text()")


def get_all_xml_files():
    xml_files = []
    for root, dirnames, filenames in os.walk('.'):
        for filename in fnmatch.filter(filenames, '*.xml'):
            xml_files.append(os.path.join(root, filename))
    return xml_files


if __name__ == '__main__':
    save_test_result_to_excel_file()
